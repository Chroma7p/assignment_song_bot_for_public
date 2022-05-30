import openpyxl
from discord.ext import commands

wb=openpyxl.load_workbook("ratingmanage/ONGEKI/ongeki.xlsx",data_only=True)
ws=wb["CRYATAL"]


rated=[[ws[j+str(i)].value for j in list("ABCDF")]for i in range(2,32)]
new=[[ws[j+str(i)].value for j in list("HIJKM")]for i in range(2,17)]
ws=wb["BP"]
bp=[[ws[j+str(i)].value for j in list("ABCF")]for i in range(2,14)]

class mycog(commands.Cog,name="製作者個人用"):
  def __init__(self,bot):
    self.bot=bot

  @commands.command(description="レーティング対象曲を表示")
  async def rated(self,ctx):
    out=""
    for i in rated:
      out+=" ".join(map(str,i))+"\n"
    await ctx.send(out)

  @commands.command(description="新曲枠を表示")
  async def newrated(self,ctx):
    out=""
    for i in new:
      out+=" ".join(map(str,i))+"\n"
    await ctx.send(out)

  @commands.command(description="BP対象曲を表示")
  async def bprated(self,ctx):
    out=""
    for i in bp:
      out+=" ".join(map(str,i))+"\n"
    await ctx.send(out)

def setup(bot):
  print("mycog setup OK")
  return bot.add_cog(mycog(bot))

