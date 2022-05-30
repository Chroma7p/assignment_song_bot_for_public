import lxml.html
from bs4 import BeautifulSoup
import openpyxl
pas=""
sco=[]
tit=[]
dif=[]
def scrap(file):
    soup=BeautifulSoup(file,'html.parser')
    titlepick=soup.find_all("form")
    ret=[]
    for youso in titlepick:
        r=[]
        r.append(youso.find("div",class_="music_label p_5 break").string)
        if youso.find("td",class_="score_label expert_score_label t_c black")!=None:
            r.append("EXPERT")
        elif youso.find("td",class_="score_label lunatic_score_label t_c black")!=None:
            r.append("LUNATIC")
        else:
            r.append("MASTER")
        
        r.append(youso.find("div",class_="f_14 l_h_12").string)
        ret.append(r)
    return ret

def difsc(file):
    soup=BeautifulSoup(file,'html.parser')
    pick=soup.find_all("div")


with open("geki.html",encoding="UTF-8")as f:
    text=f.read()
ret=scrap(text)
for i in ret:
  print(i)

wb=openpyxl.load_workbook(pas+"gekisample.xlsx")
ws=wb["RATING"]
for (i,re) in zip(range(2,32),ret[15:]):
    print(re)
    sc=""
    for j in list(re[2]):
      if j!=",":
        sc+=j
    r=str(i)
    tit=ws["A"+r]
    tit.value=re[0]
    dif=ws["B"+r]
    dif.value=re[1]
    sco=ws["D"+r]
    sco.value=int(sc)
for (i,re) in zip(range(2,17),ret[:15]):
    print(re)
    sc=""
    for j in list(re[2]):
      if j!=",":
        sc+=j
    r=str(i)
    tit=ws["H"+r]
    tit.value=re[0]
    dif=ws["I"+r]
    dif.value=re[1]
    sco=ws["K"+r]
    sco.value=int(sc)
wb.save(pas+"ongeki.xlsx")

input()
