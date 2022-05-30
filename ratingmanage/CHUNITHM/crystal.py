import lxml.html
from bs4 import BeautifulSoup
import openpyxl
pas=""
def scrap(file):
    soup=BeautifulSoup(file,'html.parser')
    titlepick=soup.find_all("form")
    ret=[]
    for youso in titlepick:
        tit=youso.find("div",class_="music_title").string
        sco=youso.find("span",class_="text_b").string
        if youso.find("div",class_="w388 musiclist_box bg_master")==None:
            dif="EXP"
        else:
            dif="MAS"

        ret.append([tit,dif,sco])
    return ret

#with open(pas+"../../csv/crys.html",encoding="UTF-8")as f:
with open("crys.html",encoding="UTF-8")as f:
  text=f.read()
ret=scrap(text)
wb=openpyxl.load_workbook(pas+"crystalsample.xlsx")
ws=wb["CRYSTAL"]
for (i,re) in zip(range(2,32),ret):
    print(re)
    sc=""
    for j in list(re[2]):
      if j!=",":
        sc+=j
    r=str(i)
    tit=ws["A"+r]
    tit.value=re[0]
    dif=ws["B"+r]
    dif.value=re[1]
    sco=ws["D"+r]
    sco.value=int(sc)
wb.save(pas+"crystal.xlsx")